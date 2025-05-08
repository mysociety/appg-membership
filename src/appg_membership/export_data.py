"""
Module for exporting APPG data to Excel spreadsheet formats for external crowdsourcing.
"""

import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from rich.console import Console

from .models import APPG, APPGList

console = Console()


def generate_google_search_link(appg_title: str) -> str:
    """
    Generate a Google search link for an APPG
    """
    search_term = f"All-Party Parliamentary Group {appg_title} UK parliament"
    encoded_search = urllib.parse.quote(search_term)
    return f"https://www.google.com/search?q={encoded_search}"


def determine_starting_status(appg: APPG) -> str:
    """
    Determine the starting status for an APPG to be put in the CSV
    """
    has_website = appg.has_website()
    has_members = (
        len(appg.members_list.members) > 0
        and appg.members_list.source_method != "empty"
    )

    if not has_website:
        return "no_website"
    elif has_website and not has_members:
        return "website_no_members"
    elif has_website and has_members:
        return "website_members_list"
    else:
        return "website"


def export_for_crowdsource(output_path: Optional[str] = None) -> str:
    """
    Export an Excel file with APPG data for external crowdsourcing.

    Fields:
    - starting_status: Current status (no_website, website, website_no_members, website_members_list)
    - review_status: Blank at export, to be filled by crowdsourcers
    - appg_slug: Unique identifier for the APPG
    - appg_name: Full name of the APPG
    - parliament_source_url: URL to the official parliament page for this APPG
    - google_link: Prepopulated Google search link for the APPG
    - appg_website: Current website URL if available
    - appg_members_page: Members page URL if available (currently same as website)

    Args:
        output_path: Optional path for the output Excel file

    Returns:
        The path to the created Excel file
    """
    appgs = APPGList.load()

    # If no output path is specified, create one with a timestamp
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"data/exports/appg_crowdsource_{timestamp}.xlsx"

    # Ensure directory exists
    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet found in the workbook.")

    ws.title = "APPG Information"

    # Define column headers
    headers = [
        "starting_status",
        "review_status",
        "appg_slug",
        "appg_name",
        "parliament_source_url",
        "google_link",
        "appg_website",
        "appg_members_page",
    ]

    # Apply header styling
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )

    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Populate data rows
    row_num = 2
    for appg in appgs:
        website_url = (
            str(appg.contact_details.website.url)
            if appg.contact_details.website.url
            else ""
        )

        # Populate cells
        ws.cell(row=row_num, column=1, value=determine_starting_status(appg))
        ws.cell(row=row_num, column=2, value="")  # review_status is blank at export
        ws.cell(row=row_num, column=3, value=appg.slug)
        ws.cell(row=row_num, column=4, value=appg.title)
        ws.cell(
            row=row_num, column=5, value=str(appg.source_url) if appg.source_url else ""
        )
        ws.cell(row=row_num, column=6, value=generate_google_search_link(appg.title))
        ws.cell(row=row_num, column=7, value=website_url)
        ws.cell(row=row_num, column=8, value=website_url)  # Initially same as website

        row_num += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)  # type: ignore
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = min(
                        len(str(cell.value)), 100
                    )  # Cap at 100 for very long URLs
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    # Save the workbook
    wb.save(output_path)

    console.print(f"[green]Excel export created at:[/green] {output_path}")
    console.print(f"[green]Exported {len(appgs)} APPGs[/green]")

    return output_path
